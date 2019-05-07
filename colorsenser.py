import RPi.GPIO as GPIO
import time
import os
import serial
import string
import pynmea2
import openpyxl 

s2 = 23
s3 = 24
signal = 25
NUM_CYCLES = 10

   


def setup():
  GPIO.setmode(GPIO.BCM)
  GPIO.setup(signal,GPIO.IN, pull_up_down=GPIO.PUD_UP)
  GPIO.setwarnings(False)
  GPIO.setup(s2,GPIO.OUT)
  GPIO.setup(s3,GPIO.OUT)
  print("\n")
  
def mapping(i,p,q,n,m):
  x=((i-q)*(n-m)/(p-q))+m
  return(x)

def redf():
    GPIO.output(s2,GPIO.LOW)
    GPIO.output(s3,GPIO.LOW)
    suma=0
    for i in range(5):
      start = time.time()
      for impulse_count in range(NUM_CYCLES):
        GPIO.wait_for_edge(signal, GPIO.FALLING)
      duration = time.time() - start      #seconds to run for loop
      suma=suma+ NUM_CYCLES / duration   #in Hz
    red=mapping(suma/5,24000,13000,255,0)
    return(red)

def bluef():
    GPIO.output(s2,GPIO.LOW)
    GPIO.output(s3,GPIO.HIGH)
    suma=0
    for i in range(5):
      start = time.time()
      for impulse_count in range(NUM_CYCLES):
        GPIO.wait_for_edge(signal, GPIO.FALLING)
      duration = time.time() - start
      suma = suma+NUM_CYCLES / duration
    blue=mapping(suma/5,27000,20000,255,0)
    return (blue)

def greenf():
    GPIO.output(s2,GPIO.HIGH)
    GPIO.output(s3,GPIO.LOW)
    suma=0;
    for i in range(5):
      start = time.time()    
      for impulse_count in range(NUM_CYCLES):
        GPIO.wait_for_edge(signal, GPIO.FALLING)
      duration = time.time() - start
      suma = suma+ NUM_CYCLES / duration
    green=mapping(suma/5,21000,16000,255,0)
    return (green)

def whitef():
    suma=(redf()+bluef()+greenf())/3
    return (mapping(suma,255,0,150,0))


def nitrogen():
    GPIO.output(s2,GPIO.HIGH)
    GPIO.output(s3,GPIO.HIGH)
    suma=0
    for i in range(5):
      start = time.time()
      for impulse_count in range(NUM_CYCLES):
        GPIO.wait_for_edge(signal, GPIO.FALLING)
      duration = time.time() - start      #seconds to run for loop
      suma=suma+ NUM_CYCLES / duration   #in Hz
    red=mapping(suma/5,37000,25000,0,-50)
    if(red>0):
        red=0
    return(red*-1)

def phosphorous():
    suma=0
    for i in range(5):
      GPIO.output(s2,GPIO.LOW)
      GPIO.output(s3,GPIO.LOW)
      start = time.time()
      for impulse_count in range(NUM_CYCLES):
        GPIO.wait_for_edge(signal, GPIO.FALLING)
      duration = time.time() - start
      if(i!=0):
          suma = suma+NUM_CYCLES / duration
      print(NUM_CYCLES/duration)
    print(suma/5)
    blue=mapping(suma/5,24000,19000,0,-45)
    return (blue*-1)


def ph():
    r=redf()
    g=greenf()
    b=blue()
    if(r>204):      #red high
        if(g>204):
            return 4
        elif(g>153):
            return 3
        elif(g>102):
            return 2
        elif(g>51):
            return 1
        else:      
            if(b>127):
                return 14
            else:
                return int(0)
    elif(r<=51):        #red low
        if(g<=85):
             return 10
        elif(g<=170):
            return 9
        else:
            if(b>127):
                return 8
            else:
                return 7
    else:                   #red in mid value
        if(b>127):
            if(r>153):
                return 13
            elif(r>102):
                return 12
            else:
                return 11
        else:
            if(r>127):
                return 5
            else:
                return 6

def loop():
    nit=nitrogen()
    print("The test result of nitrogen is ",nit)
    input()
    pho=phosphorous()
    print("\n\n The result of Phosphorous is",pho)
    input()
    pot=whitef()
    print("\n\n The result of Potassium is ",pot)
    input()
    phf=ph()
    print("\n\n The value of ph is ",phf)
    input()
    
    port = "/dev/ttyAMA0"
    ser= serial.Serial(port,baudrate=9600,timeout=0.5)
    dataout= pynmea2.NMEAStreamReader()
    newdata = ser.readline()
    print("GEt lat and lon")
    if newdata[0:6] =='$GPGGA':
        newmsg= pynmea2.parse(newdata)
        lat =newmsg.latitude
        print(lat)
        la=newmsg.longitude
        print(la)
        time.sleep(10)
    input()

    
    print("\n\n Transfering data to tsl.xlsx")
    input()
    wb = openpyxl.load_workbook('/home/pi/ts1.xlsx') 
    sheet = wb.active 
    m_row=sheet.max_row
    cs= sheet.cell(row=max_row,column=1)
    cs.value=nit
    cs= sheet.cell(row=max_row,column=2)
    cs.value=pho
    cs= sheet.cell(row=max_row,column=3)
    cs.value=pot
    cs= sheet.cell(row=max_row,column=4)
    cs.value=phf
    cs= sheet.cell(row=max_row,column=5)
    cs.value=lat
    cs= sheet.cell(row=max_row,column=6)
    cs.value=la
    wb.save('/home/pi/ts1.xlsx') 

    

def endprogram():
    GPIO.cleanup()

if __name__=='__main__':
    
    setup()
    try:
        loop()
    except KeyboardInterrupt:
        endprogram()

